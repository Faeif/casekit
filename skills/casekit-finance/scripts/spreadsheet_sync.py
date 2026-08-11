#!/usr/bin/env python3
"""Inspect Excel/CSV data and synchronise mapped values into CaseKit's metric tree."""

import argparse
import csv
import json
from datetime import datetime, timezone
from pathlib import Path


def number(value, context):
    if isinstance(value, bool) or value is None:
        raise ValueError(f"{context}: expected numeric value, got {value!r}")
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).replace(",", "").strip())
    except ValueError as exc:
        raise ValueError(f"{context}: expected numeric value, got {value!r}") from exc


def format_number(value):
    if value.is_integer():
        return str(int(value))
    return f"{value:.12f}".rstrip("0").rstrip(".")


def read_csv(path):
    with path.open(newline="", encoding="utf-8-sig") as handle:
        rows = list(csv.reader(handle, delimiter="\t" if path.suffix.lower() == ".tsv" else ","))
    return {"CSV": rows}


def read_xlsx(path, values_only=False):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise SystemExit("openpyxl is required for .xlsx. Run: python -m pip install -r requirements.txt") from exc
    book = load_workbook(path, data_only=values_only, read_only=True)
    result = {}
    for sheet in book.worksheets:
        result[sheet.title] = [[cell.value for cell in row] for row in sheet.iter_rows()]
    return result


def workbook(path, values_only=False):
    suffix = path.suffix.lower()
    if suffix in {".csv", ".tsv"}:
        return read_csv(path)
    if suffix == ".xlsx":
        return read_xlsx(path, values_only)
    raise ValueError("Supported spreadsheet formats: .xlsx, .csv, .tsv")


def cell_value(path, sheet_name, cell):
    if path.suffix.lower() != ".xlsx":
        raise ValueError("Cell mappings require .xlsx inputs; import CSV values into the metric tree directly.")
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise SystemExit("openpyxl is required for .xlsx. Run: python -m pip install -r requirements.txt") from exc
    formula_book = load_workbook(path, data_only=False, read_only=True)
    value_book = load_workbook(path, data_only=True, read_only=True)
    if sheet_name not in formula_book.sheetnames:
        raise ValueError(f"Sheet not found: {sheet_name}")
    formula = formula_book[sheet_name][cell].value
    value = value_book[sheet_name][cell].value
    if isinstance(formula, str) and formula.startswith("=") and value is None:
        raise ValueError(f"{sheet_name}!{cell} has a formula without a cached result. Recalculate and save the workbook in Excel first.")
    return value, formula


def markdown_report(path):
    raw = workbook(path, values_only=False)
    values = workbook(path, values_only=True) if path.suffix.lower() == ".xlsx" else raw
    lines = [f"# Spreadsheet inspection: {path.name}", "", f"- Path: `{path}`", "- Values from formulas use Excel's last saved calculation cache.", ""]
    for name, rows in raw.items():
        nonempty = [row for row in rows if any(value not in (None, "") for value in row)]
        formulas = sum(1 for row in rows for value in row if isinstance(value, str) and value.startswith("="))
        lines.extend([f"## {name}", "", f"- Non-empty rows: {len(nonempty)}", f"- Formula cells: {formulas}", "", "### Preview", ""])
        preview = values[name][: min(12, len(values[name]))]
        for row in preview:
            lines.append(" | ".join("" if value is None else str(value) for value in row[:12]))
        lines.append("")
    return "\n".join(lines)


def read_metric_tree(path):
    with path.open(newline="", encoding="utf-8-sig") as handle:
        reader = csv.DictReader(handle)
        return reader.fieldnames or [], list(reader)


def write_metric_tree(path, fields, rows):
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        writer.writerows(rows)


def official_dir(project):
    clean = project / "03-OFFICIAL"
    return clean if clean.is_dir() else project


def sync(project, mapping_path, apply):
    mapping = json.loads(mapping_path.read_text(encoding="utf-8"))
    mappings = mapping.get("mappings")
    if not isinstance(mappings, list) or not mappings:
        raise ValueError("Mapping must contain a non-empty 'mappings' list")
    metric_path = official_dir(project) / "03-metric-tree.csv"
    fields, rows = read_metric_tree(metric_path)
    required = {"metric_id", "low", "base", "high"}
    if not required <= set(fields):
        raise ValueError("03-metric-tree.csv is missing required scenario columns")
    by_id = {row.get("metric_id"): row for row in rows}
    report = {"generated_at": datetime.now(timezone.utc).isoformat(), "mapping": str(mapping_path), "updates": [], "warnings": []}
    for item in mappings:
        for key in ("metric_id", "scenario", "file", "sheet", "cell"):
            if key not in item:
                raise ValueError(f"Mapping item missing '{key}'")
        scenario = item["scenario"]
        if scenario not in {"low", "base", "high"}:
            raise ValueError(f"Unsupported scenario '{scenario}'")
        metric_id = item["metric_id"]
        if metric_id not in by_id:
            raise ValueError(f"Mapping references unknown metric {metric_id}")
        source = (project / item["file"]).resolve()
        if not source.is_file():
            raise ValueError(f"Mapped spreadsheet does not exist: {source}")
        value, formula = cell_value(source, item["sheet"], item["cell"])
        parsed = number(value, f"{source.name}:{item['sheet']}!{item['cell']}")
        old = by_id[metric_id].get(scenario, "")
        report["updates"].append({"metric_id": metric_id, "scenario": scenario, "old": old, "new": parsed, "file": item["file"], "cell": f"{item['sheet']}!{item['cell']}", "formula": formula if isinstance(formula, str) and formula.startswith("=") else None})
        by_id[metric_id][scenario] = format_number(parsed)
    if apply:
        write_metric_tree(metric_path, fields, rows)
    return report


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    sub = parser.add_subparsers(dest="command", required=True)
    inspect = sub.add_parser("inspect")
    inspect.add_argument("file", type=Path)
    inspect.add_argument("--output", type=Path)
    sync_parser = sub.add_parser("sync")
    sync_parser.add_argument("project", type=Path)
    sync_parser.add_argument("mapping", type=Path)
    sync_parser.add_argument("--apply", action="store_true")
    sync_parser.add_argument("--report", type=Path)
    args = parser.parse_args()
    if args.command == "inspect":
        report = markdown_report(args.file.expanduser().resolve())
        if args.output:
            output = args.output.expanduser().resolve()
            output.parent.mkdir(parents=True, exist_ok=True)
            output.write_text(report, encoding="utf-8")
            print(f"Wrote workbook inspection -> {output}")
        else:
            print(report)
        return
    report = sync(args.project.expanduser().resolve(), args.mapping.expanduser().resolve(), args.apply)
    if args.report:
        args.report.expanduser().resolve().parent.mkdir(parents=True, exist_ok=True)
        args.report.expanduser().resolve().write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(report, ensure_ascii=False, indent=2))
    if not args.apply:
        print("Preview only. Run again with --apply to update 03-metric-tree.csv.")


if __name__ == "__main__":
    main()
